# iinicio em 01/12/2022 12:00
# termino em 02/12/2022 11:30
# https://github.com/Samuel-Batista/excelToWordTable


# importando blibioteca que cuida de arquivos .xlsx(excel)
from openpyxl import load_workbook

# importando blibioteca que cuida de edição de .doc(word)
from docx import Document


# variaveis de referencia
word_template = Document('doc.docx')
word_template_table = word_template.tables[1]
excel = load_workbook(filename='excel.xlsx')
lista_pedidos = excel['PT']
listagem = excel['Listagem']


# variaveis para controlhe do loop
count = 1
current_column = 1
current_row = 4
current_cell_codigo_produto = lista_pedidos.cell(column=current_column, row=current_row)


def pegar_nome_em_listagem(codigo):
    current_row = 2
    current_column = 1
    current_cell_codigo = listagem.cell(column=current_column, row=current_row)

    # loop continua quando a celula tem valor
    while not current_cell_codigo.value == None:
        
        # se o codigo de listagem for igual ao codigo passado pela função: retorna o nome
        if current_cell_codigo.value == codigo:
            return listagem.cell(column=2, row=current_row).value
             
        # aumentar linha
        current_row += 1

        # atualizar o valor de current_cell_codigo
        current_cell_codigo = listagem.cell(column=current_column, row=current_row)

    return None


# verificar cada linha da primeira coluna da lista de pedidos
while current_cell_codigo_produto.value:
    if count > 26:
        break

    current_name = pegar_nome_em_listagem(current_cell_codigo_produto.value)
    current_code = current_cell_codigo_produto.value

    # inserir codigo no word
    word_template_table.cell(count, 0).text = str(current_code)

    # inserir nome no word
    word_template_table.cell(count, 2).text = current_name

    count += 1
    current_row += 1
    current_cell_codigo_produto = lista_pedidos.cell(column=current_column, row=current_row)



word_template.save("result.docx")